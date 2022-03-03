import win32com.client as win32
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill,Font

import os

def add_if_key_not_exist(dict_obj, key, value):
    """ Add new key-value pair to dictionary only if
    key does not exist in dictionary. """
    if key not in dict_obj:
        dict_obj.update({key: value})


selected=filedialog.askopenfilename()
selected=selected.replace("/","\\")
new_name=selected+"x" #change to xlsm


#======xls trans to xlsx in order to use openpyxl
excel=win32.gencache.EnsureDispatch('Excel.Application')
wb=excel.Workbooks.Open(selected)
wb.SaveAs(new_name,FileFormat=51)
wb.Close()
excel.Application.Quit()

#====read WHCK report comments==================
wb=load_workbook(new_name)
head,tail=os.path.split(new_name)
tail="[new]"+tail
new_excel=head+tail
wb.active=wb['WHCK Report']
ws=wb.active
max_rows=ws.max_row


#find the empty column to define how many columns.
empty_col=ws.max_column

#print(empty_col)
empty_fill = PatternFill(fill_type='solid',start_color='FFFFFF',end_color='FFFFFF')
pass_fill = PatternFill(fill_type='solid',start_color='00BB00',end_color='00BB00')
fail_fill = PatternFill(fill_type='solid',start_color='FF0000',end_color='FF0000')
notrun_fill = PatternFill(fill_type='solid',start_color='ADADAD',end_color='ADADAD')
cali = Font(name='Calibri')

for x in range(12,max_rows):
	d={}
	ws.cell(row=x,column=1).fill = empty_fill
	ws.cell(row=x,column=1).font = cali


	for y in range(2,empty_col):
		
		if ws.cell(row=x,column=y).value =="Passed":
			ws.cell(row=x,column=y).fill = pass_fill
		if ws.cell(row=x,column=y).value =="NotRun":
			ws.cell(row=x,column=y).fill = notrun_fill
		if ws.cell(row=x,column=y).value =="Failed":
			ws.cell(row=x,column=y).fill = fail_fill
		if ws.cell(row=x,column=y).comment is not None:
			tmp=ws.cell(row=x,column=y).comment.text
			tmp=tmp.split("\n")
			f_id=[]
			for k in tmp:
				if "Filter" in k:
					new=k.split(" ")
					new=new[0].replace("Filter","")
					f_id.append(new)
					add_if_key_not_exist(d,new,0)
		filter_id=""
			# if len(f_id)==1:
			# 	filter_id=f_id[0]
			# else:
			# 	for p in range(0,len(f_id)):
			# 		filter_id+=f_id[p]+"/"
			# 	filter_id=filter_id[0:-1]
		for m in d.keys():
			filter_id+=m+"/"
		filter_id = filter_id[0:-1]
			#print(d)
			# print(filter_id)
			# print(filter_id)
			
		ws.cell(row=x,column=empty_col-1).value=filter_id
		ws.cell(row=x,column=empty_col-1).fill = empty_fill
		ws.cell(row=x,column=y).font = cali
		ws.cell(row=x,column=empty_col-1).font = cali

#==========filter_items__read===============

wb.active=wb['Filter Summary']
ws=wb.active
max_rows_2=ws.max_row
# ID=[]
# des=[]
# date=[]
# test_name=[]
all_info=[]
# for x in range(1,max_rows_2):
# 	cell=ws.cell(row=x,column=3)
# 	test_name.append(cell.value)

# for x in range(1,max_rows_2):
# 	cell=ws.cell(row=x,column=4)
# 	ID.append(cell.value)

# for x in range(1,max_rows_2):
# 	cell=ws.cell(row=x,column=5)
# 	date.append(cell.value)

# for x in range(1,max_rows_2):
# 	cell=ws.cell(row=x,column=6)
# 	des.append(cell.value)

for x in range(1,max_rows_2+1):
	a=[]
	ID=ws.cell(row=x,column=4) #ID
	a.append(ID.value)
	des=ws.cell(row=x,column=6) #des
	a.append(des.value)
	date=ws.cell(row=x,column=5) #date
	a.append(date.value)
	test_name=ws.cell(row=x,column=3) #test name
	a.append(test_name.value)
	if (a not in all_info):
		all_info.append(a)


#==========filter_items__write===============
wb.create_sheet("Dell_Filter")
wb.active=wb['Dell_Filter']
ws=wb.active

# for x in range(1,max_rows_2):
# 	ws.cell(row=x,column=1).value=ID[x-1]

# for x in range(1,max_rows_2):
# 	ws.cell(row=x,column=2).value=des[x-1]

# for x in range(1,max_rows_2):
# 	ws.cell(row=x,column=3).value=date[x-1]

# for x in range(1,max_rows_2):
# 	ws.cell(row=x,column=4).value=test_name[x-1]

for x in range(1,len(all_info)+1):
	for y in range(1,5):
		ws.cell(row=x,column=y).value=all_info[x-1][y-1]
		ws.cell(row=x,column=y).font = cali

wb.remove(wb['Filter Summary'])



wb.save(new_name)


